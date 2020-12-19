from flask import Flask, request
import openpyxl
import json
import datetime


sheet=None
boo=None
s=[]
rown=1
app = Flask(__name__)
 
@app.route('/', methods=['POST', 'GET'])
def index():
        global s,rown,sheet,boo
        if request.method == 'POST':
                check=request.get_json()
                for item in check["check"]:
                        cel = [check["user_id"],item["item"],item["price"]]
                        s.append(cel)
                if len(s)>1000:
                        try:
                                f=open("book.xlsx")
                        except FileNotFoundError:
                                        boo=openpyxl.Workbook()
                                        sheet=boo.active
                                        boo.save("book.xlsx")
                                        boo.close()
                                        f=open("book.xlsx")
                        if f:
                                boo=openpyxl.open("book.xlsx")
                                sheet=boo.active
                                if sheet.cell(row=1,column=1).value is None:
                                        sheet.cell(rown,column=1).value='N'
                                        sheet.cell(rown,column=2).value='User ID'
                                        sheet.cell(rown,column=3).value='Datetime'
                                        sheet.cell(rown,column=4).value='Item'
                                        sheet.cell(rown,column=5).value='Prise'
                                while sheet.cell(rown,column=1).value is not None:
                                        rown+=1
                                for cel in s:
                                        for it in range(1):
                                                d1=datetime.datetime.today()
                                                d1=d1.strftime("%a %d %b %X")
                                                sheet[rown][0].value=rown-1
                                                sheet[rown][1].value=cel[it]
                                                sheet[rown][2].value=d1
                                                sheet[rown][3].value=cel[it+1]
                                                sheet[rown][4].value=cel[it+2]
                                        rown+=1     
                                boo.save("book.xlsx")
                                boo.close()
                        s.clear()
                                
                        
        return 'OK'
if __name__ == "__main__":
    app.run()